import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Optional, TypedDict, Union


class Data(TypedDict):
    text: str
    url: Optional[str]


class ExcelHelper:
    ws: Worksheet
    wb: Workbook

    def __init__(self, downloadDir: str, fileName: str, sheetName: str = "Sheet1"):
        self.path = os.path.join(downloadDir, fileName)
        print(f"엑셀 파일명: {fileName}", flush=True)
        if os.path.exists(self.path):
            self.wb = load_workbook(self.path)
            if sheetName in self.wb.sheetnames:
                self.ws = self.wb[sheetName]
            else:
                self.ws = self.wb.create_sheet(sheetName)
            print(f"Excel 데이터 로드 완료, {self.path}", flush=True)
        else:
            self.wb = Workbook()
            activeWs = self.wb.active
            assert isinstance(activeWs, Worksheet)
            self.ws = activeWs
            self.ws.title = sheetName
            self.setData(
                ["검색어", "기관명", "정보 제목", "단위 업무", "생산 일자", "파일 링크"]
            )
            print(f"Excel 데이터 생성 완료, {self.path}", flush=True)

    def setData(self, datas: List[Union[Data, str]]):
        nextRow = self.ws.max_row + 1

        for idx, data in enumerate(datas, start=0):
            cell = self.ws.cell(row=nextRow, column=idx + 1)
            if isinstance(data, str):
                text, url = data, None
            else:
                text, url = data["text"], data["url"]

            cell.value = text  # type: ignore
            if url:
                cell.hyperlink = url  # type: ignore
                cell.style = "Hyperlink"

            print(f"{nextRow}_{idx + 1}에 {text} 삽입 완료", flush=True)

    def notFoundData(self, query, organization, text):
        nextRow = self.ws.max_row + 1
        data = [query, organization, text]
        self.ws.append(data)
        print(f"{nextRow}에 {query}-{organization}-{text} 삽입 완료", flush=True)

    def setHyperlink(
        self,
        fileLinks: List,
        col: int,
        displayText: str = "바로가기",
        hasMissingDownloads: bool = False,
    ):
        row = self.ws.max_row
        missingFile = False

        for idx, link in enumerate(fileLinks, start=0):
            if os.path.exists(link):
                path = f"file:///{os.path.abspath(link)}"
                cell = self.ws.cell(row=row, column=col + idx)
                cell.value = displayText  # type: ignore
                cell.hyperlink = path  # type: ignore
                cell.style = "Hyperlink"
                print(
                    f"Excel에 {os.path.abspath(link)} 파일 연결 완료",
                    flush=True,
                )
            else:
                print(f"링크 대상 파일을 찾을 수 없습니다: {link}", flush=True)
                missingFile = True

        if hasMissingDownloads or missingFile:
            warnCell = self.ws.cell(row=row, column=col + len(fileLinks))
            warnCell.value = "누락된 파일이 존재합니다."  # type: ignore
            warnCell.font = Font(color="FFFF0000")

    def save(self):
        directory = os.path.dirname(self.path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
        self.wb.save(self.path)
        print(f"{directory}에 Excel 저장 완료", flush=True)

    def pretterColumns(self):
        ws = self.ws
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter  # type: ignore
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)  # type: ignore
                except:
                    pass
            adjusted_width = (max_length + 2) * 2
            ws.column_dimensions[column_letter].width = adjusted_width
